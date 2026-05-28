#!/usr/bin/env python3
"""Scrape le taux de remplissage des seances Re-SET (booking bsport).

Recupere, jour par jour et seance par seance, le nombre de personnes
presentes / la capacite de chaque seance, et genere :
  - reset_seances.csv   (donnees brutes)
  - reset_seances.xlsx  (Excel mis en forme)
  - index.html          (dashboard autonome, avec bouton de mise a jour)

Relancer le script reactualise tout. Le dashboard publie peut aussi se
mettre a jour tout seul (bouton "Mettre a jour") en appelant directement
l'API bsport depuis le navigateur (de la derniere date connue jusqu'a hier).

Usage:
    python3 reset_scrape.py                          # 22/03/2026 -> hier
    python3 reset_scrape.py --start 2026-03-22
    python3 reset_scrape.py --start 2026-03-22 --end 2026-06-30

Source: https://www.re-set.club/reservation  (widget bsport, company 5181)
"""
import argparse
import csv
import datetime as dt
import json
import os
import sys
import time
import urllib.parse
import urllib.request

API = "https://api.production.bsport.io"
COMPANY = 5181
JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def _get(path, params, retries=4):
    url = f"{API}{path}?" + urllib.parse.urlencode(params)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(2 ** attempt)
    raise RuntimeError(f"echec requete {url}: {last}")


def fetch_coaches():
    """Table associated_coach_id -> nom complet (best effort)."""
    mapping = {}
    try:
        data = _get("/book/v1/associated_coach/", {"company": COMPANY, "page_size": 500})
        for c in data.get("results", []):
            name = c.get("name") or f"{c.get('firstname','')} {c.get('lastname','')}".strip()
            for key in ("associated_coach_id", "id"):
                if c.get(key):
                    mapping[c[key]] = name
            for aid in c.get("associatedcoach_set", []) or []:
                mapping.setdefault(aid, name)
    except Exception as e:  # noqa: BLE001
        print(f"  (avertissement: noms des coachs indisponibles: {e})", file=sys.stderr)
    return mapping


def fetch_offers(start, end):
    """Toutes les seances entre start et end (dates incluses)."""
    offers = []
    page = 1
    while True:
        data = _get("/book/v1/offer/", {
            "company": COMPANY,
            "only_future_strict": "false",
            "min_date": start.isoformat(),
            "max_date": end.isoformat(),
            "page_size": 300,
            "page": page,
        })
        results = data.get("results", [])
        offers.extend(results)
        if not data.get("next_page") or not results:
            break
        page += 1
    return offers


def build_rows(offers, coaches):
    """Transforme les offres en lignes, dedupliquees par session_id."""
    by_id = {}
    for o in offers:
        sid = o.get("id")
        ds = o.get("date_start", "") or ""
        local = ds[:16].replace("T", " ")  # 'AAAA-MM-JJ HH:MM' heure locale Paris
        date_part = local[:10]
        heure = local[11:16]
        try:
            jour = JOURS_FR[dt.date.fromisoformat(date_part).weekday()]
        except ValueError:
            jour = ""
        present = o.get("validated_booking_count") or 0
        capacite = o.get("effectif") or 0
        taux = round(100 * present / capacite, 1) if capacite else ""
        coach_id = o.get("coach")
        by_id[sid] = {
            "date": date_part,
            "jour": jour,
            "heure": heure,
            "activite": o.get("activity_name", "") or "",
            "coach": coaches.get(coach_id, coach_id if coach_id else ""),
            "presents": present,
            "capacite": capacite,
            "remplissage": f"{present}/{capacite}",
            "taux_%": taux,
            "complet": "oui" if o.get("full") else "non",
            "session_id": sid,
        }
    rows = list(by_id.values())
    rows.sort(key=lambda r: (r["date"], r["heure"]))
    return rows


FIELDS = ["date", "jour", "heure", "activite", "coach", "presents",
          "capacite", "remplissage", "taux_%", "complet", "session_id"]


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    print(f"-> {path}")


def write_xlsx(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (Excel ignore: 'pip install openpyxl' pour l'activer)", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Seances"
    head_fill = PatternFill("solid", fgColor="3A211B")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append([h.replace("_", " ").capitalize() for h in FIELDS])
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")
    date_col = FIELDS.index("date") + 1
    for r in rows:
        ws.append([r[h] for h in FIELDS])
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        cell = row[0]
        try:
            cell.value = dt.date.fromisoformat(cell.value)
            cell.number_format = "DD/MM/YYYY"
        except (ValueError, TypeError):
            pass
    taux_col = FIELDS.index("taux_%") + 1
    for row in ws.iter_rows(min_row=2, min_col=taux_col, max_col=taux_col):
        cell = row[0]
        v = cell.value
        if isinstance(v, (int, float)):
            if v >= 85:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif v >= 50:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
            else:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
    widths = [12, 11, 7, 18, 26, 9, 9, 12, 9, 9, 12]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"-> {path}")


def write_html(rows, coaches, path, start, end):
    chartjs = ""
    vendor = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor_chartjs.min.js")
    if os.path.exists(vendor):
        with open(vendor, encoding="utf-8") as f:
            chartjs = f.read()
    else:
        chartjs = ('document.write(\'<scr\'+\'ipt src="https://cdn.jsdelivr.net/npm/'
                   'chart.js@4.4.1/dist/chart.umd.min.js"><\\/scr\'+\'ipt>\');')
    html = (HTML_TEMPLATE
            .replace("__CHARTJS__", chartjs)
            .replace("__DATA__", json.dumps(rows, ensure_ascii=False))
            .replace("__COACHES__", json.dumps(coaches, ensure_ascii=False))
            .replace("__COMPANY__", str(COMPANY))
            .replace("__BUILDTS__", dt.datetime.now().strftime("%Y%m%d%H%M%S"))
            .replace("__GENERATED__", dt.datetime.now().strftime("%d/%m/%Y %H:%M"))
            .replace("__PERIODE__", f"{start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"))
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"-> {path}")


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Re-SET - Fréquentation des séances</title>
<script>__CHARTJS__</script>
<style>
  :root{--bg:#241410;--card:#33201a;--card2:#3d271f;--line:#4d342a;
        --text:#f3e7df;--muted:#bda394;--accent:#d98b63;--accent2:#e8b08a;
        --green:#7bbf86;--yellow:#e0c06a;--red:#d9776f;}
  *{box-sizing:border-box;}
  body{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
       background:var(--bg);color:var(--text);}
  header{padding:28px 32px 12px;}
  h1{margin:0;font-size:24px;letter-spacing:.5px;font-weight:700;}
  .sub{color:var(--muted);font-size:13px;margin-top:6px;}
  .wrap{padding:0 32px 48px;max-width:1280px;margin:0 auto;}
  .updbar{display:flex;flex-wrap:wrap;align-items:center;gap:12px;margin:20px 0 4px;
          background:var(--card);border:1px solid var(--line);border-radius:14px;padding:14px 18px;}
  .updbar .info{color:var(--muted);font-size:13px;}
  .updbar .info b{color:var(--text);}
  #updMsg{font-size:13px;color:var(--accent2);}
  .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:16px;margin:24px 0;}
  .kpi{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:18px 20px;}
  .kpi .v{font-size:28px;font-weight:700;color:var(--accent2);}
  .kpi .l{color:var(--muted);font-size:12px;margin-top:4px;text-transform:uppercase;letter-spacing:.6px;}
  .grid{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px;}
  @media(max-width:880px){.grid{grid-template-columns:1fr;}}
  .panel{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:18px 20px;}
  .panel h2{margin:0 0 14px;font-size:14px;color:var(--accent2);font-weight:600;letter-spacing:.4px;}
  canvas{max-height:260px;}
  .filters{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin:8px 0 16px;}
  .filters input,.filters select{background:var(--card2);color:var(--text);border:1px solid var(--line);
       border-radius:9px;padding:9px 12px;font-size:13px;}
  .filters input{flex:1;min-width:180px;}
  table{width:100%;border-collapse:collapse;font-size:13px;}
  th,td{padding:9px 10px;text-align:left;border-bottom:1px solid var(--line);white-space:nowrap;}
  th{color:var(--muted);font-weight:600;cursor:pointer;user-select:none;position:sticky;top:0;background:var(--card);}
  th:hover{color:var(--accent2);}
  tbody tr:hover{background:var(--card2);}
  .bar{display:inline-block;height:7px;border-radius:4px;vertical-align:middle;margin-right:8px;width:60px;
       background:var(--line);position:relative;overflow:hidden;}
  .bar>span{position:absolute;left:0;top:0;bottom:0;border-radius:4px;}
  .pill{padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600;}
  .tablewrap{max-height:560px;overflow:auto;border:1px solid var(--line);border-radius:14px;}
  .foot{color:var(--muted);font-size:12px;margin-top:18px;}
  a{color:var(--accent2);}
  label.fld{color:var(--muted);font-size:13px;display:inline-flex;align-items:center;}
  .seg button{background:var(--card2);color:var(--muted);border:1px solid var(--line);
       padding:8px 16px;font-size:13px;cursor:pointer;}
  .seg button:first-child{border-radius:9px 0 0 9px;}
  .seg button:last-child{border-radius:0 9px 9px 0;}
  .seg button:not(:last-child){border-right:none;}
  .seg button.on{background:var(--accent);color:#241410;font-weight:700;}
  .btn{background:var(--accent);color:#241410;border:none;border-radius:9px;padding:9px 16px;
       font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap;}
  .btn:hover{background:var(--accent2);}
  .btn:disabled{opacity:.6;cursor:default;}
  .kpi .delta{margin-top:6px;font-size:13px;font-weight:700;}
  .ranklist{display:flex;flex-direction:column;gap:9px;}
  .rk{display:grid;grid-template-columns:62px 1fr auto;align-items:center;gap:10px;font-size:13px;}
  .rk .lbl{color:var(--text);font-weight:600;}
  .rk .track{height:9px;background:var(--line);border-radius:5px;overflow:hidden;}
  .rk .track>span{display:block;height:100%;background:var(--accent);border-radius:5px;}
  .rk .val{color:var(--muted);font-variant-numeric:tabular-nums;}
  @media(max-width:600px){
    header{padding:18px 14px 6px;}
    h1{font-size:18px;line-height:1.25;}
    .sub{font-size:12px;}
    .wrap{padding:0 12px 32px;}
    .updbar{padding:12px 14px;}
    .kpis{grid-template-columns:1fr 1fr;gap:10px;margin:14px 0;}
    .kpi{padding:13px 14px;border-radius:12px;}
    .kpi .v{font-size:21px;}
    .kpi .l{font-size:10px;}
    .panel{padding:14px 13px;}
    .panel h2{font-size:13px;}
    .seg{display:flex;width:100%;}
    .seg button{flex:1;padding:10px 4px;}
    .filters{gap:8px;}
    .filters input,.filters select,.btn{font-size:15px;width:100%;}
    .filters input{min-width:0;}
    .updbar .btn{width:100%;}
    canvas{max-height:240px;}
  }
</style>
</head>
<body>
<header>
  <h1>Re-SET &middot; Fréquentation des séances</h1>
  <div class="sub">Période __PERIODE__ &middot; généré le __GENERATED__</div>
</header>
<div class="wrap">
  <div class="updbar">
    <span class="info" id="lastScrap"></span>
    <button id="btnUpdate" class="btn">Mettre à jour</button>
    <span id="updMsg"></span>
  </div>
  <div class="kpis" id="kpis"></div>
  <div class="panel">
    <h2>Tendance de la fréquentation</h2>
    <div class="kpis" id="trend"></div>
  </div>
  <div class="panel" id="caPanel">
    <h2>Chiffre d'affaires estimé</h2>
    <div class="filters">
      <label class="fld">Prix moyen par séance (&euro;)
        <input id="prix" type="number" min="0" step="0.5" value="25" style="width:90px;margin-left:8px">
      </label>
      <span class="seg" id="caSeg">
        <button data-g="jour">Par jour</button>
        <button data-g="semaine">Par semaine</button>
        <button data-g="mois" class="on">Par mois</button>
      </span>
    </div>
    <div class="kpis" id="caKpis" style="margin:6px 0 18px"></div>
    <canvas id="cCA"></canvas>
  </div>
  <div class="panel">
    <h2>Fréquentation depuis l'ouverture</h2>
    <div class="filters">
      <span class="seg" id="metSeg">
        <button data-m="visiteurs" class="on">Visiteurs</button>
        <button data-m="seances">Nb de séances</button>
      </span>
      <span class="seg" id="evSeg">
        <button data-g="jour">Jour</button>
        <button data-g="semaine">Semaine</button>
        <button data-g="mois" class="on">Mois</button>
      </span>
      <select id="evAct"></select>
    </div>
    <canvas id="cDay" style="margin-top:6px"></canvas>
  </div>
  <div class="grid">
    <div class="panel"><h2>Top créneaux horaires (visiteurs)</h2><div id="topHour" class="ranklist"></div></div>
    <div class="panel"><h2>Visiteurs par jour de la semaine</h2><div id="topDay" class="ranklist"></div></div>
  </div>
  <div class="grid">
    <div class="panel"><h2>Visiteurs par type de séance</h2><canvas id="cAct"></canvas></div>
    <div class="panel"><h2>Visiteurs par créneau horaire</h2><canvas id="cHour"></canvas></div>
  </div>
  <div class="panel">
    <h2>Visiteurs moyens par séance &amp; par coach (les stars)</h2>
    <canvas id="cCoach" style="max-height:none"></canvas>
  </div>
  <div class="panel">
    <h2>Détail des séances</h2>
    <div class="filters">
      <input id="q" placeholder="Rechercher (activité, coach, date...)">
      <select id="fAct"></select>
      <select id="fDay"></select>
      <button id="btnExport" class="btn">Exporter en Excel</button>
    </div>
    <div class="tablewrap"><table id="tbl"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="foot">Visiteurs = personnes présentes (réservations validées). Source : re-set.club (widget bsport).</div>
</div>
<script>
const API="https://api.production.bsport.io", COMPANY=__COMPANY__, BUILD_TS="__BUILDTS__";
const COACHES=__COACHES__;
const JOURS=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"];
const MOIS=['janv.','févr.','mars','avr.','mai','juin','juil.','août','sept.','oct.','nov.','déc.'];

let DATA=__DATA__;
// Reprend les mises a jour incrementales locales tant que le build n'a pas change
try{
  const s=JSON.parse(localStorage.getItem('reset_dash')||'null');
  if(s && s.buildTs===BUILD_TS && Array.isArray(s.data) && s.data.length>=DATA.length) DATA=s.data;
}catch(e){}
function persist(){try{localStorage.setItem('reset_dash',JSON.stringify({buildTs:BUILD_TS,data:DATA}));}catch(e){}}

const col=t=>t>=85?'#7bbf86':t>=50?'#e0c06a':'#d9776f';
const eur=v=>v.toLocaleString('fr-FR',{maximumFractionDigits:0})+' €';
Chart.defaults.color='#bda394';Chart.defaults.borderColor='#4d342a';
Chart.defaults.font.family=getComputedStyle(document.body).fontFamily;

function avg(arr){let p=0,c=0;arr.forEach(r=>{p+=r.presents;c+=r.capacite;});return c?p/c*100:0;}
function group(key){const m={};DATA.forEach(r=>{(m[key(r)]=m[key(r)]||[]).push(r);});return m;}
function lundi(s){const d=new Date(s+'T00:00:00');const j=(d.getDay()+6)%7;d.setDate(d.getDate()-j);return d.toISOString().slice(0,10);}
function periodKey(r,g){return g==='jour'?r.date:g==='mois'?r.date.slice(0,7):lundi(r.date);}
function fmtJ(iso){const p=iso.split('-');return `${p[2]}/${p[1]}/${p[0]}`;}
function fmtJM(iso){const p=iso.split('-');return `${p[2]}/${p[1]}`;}
function fmtMois(ym){const p=ym.split('-');return `${MOIS[+p[1]-1]} ${p[0]}`;}
function labelPeriode(k,g){return g==='mois'?fmtMois(k):g==='semaine'?'sem. '+fmtJM(k):fmtJM(k);}

let evChart=null,caChart=null,actChart=null,hourChart=null,coachChart=null;
let evGran='mois',evMetric='visiteurs',caGran='mois',sortKey='date',sortDir=1,currentRows=[];

const nf=v=>Math.round(v).toLocaleString('fr-FR');
function renderKpis(){
  const totP=DATA.reduce((s,r)=>s+r.presents,0);
  const byD={};DATA.forEach(r=>byD[r.date]=(byD[r.date]||0)+r.presents);
  const bd=Object.entries(byD).sort((a,b)=>b[1]-a[1])[0]||['',0];
  const byW={};DATA.forEach(r=>{const k=lundi(r.date);byW[k]=(byW[k]||0)+r.presents;});
  const bw=Object.entries(byW).sort((a,b)=>b[1]-a[1])[0]||['',0];
  document.getElementById('kpis').innerHTML=[
    ['Visiteurs (total)',nf(totP),''],
    ['Séances',nf(DATA.length),''],
    ['Meilleure journée',nf(bd[1]),bd[0]?fmtJ(bd[0]):''],
    ['Meilleure semaine',nf(bw[1]),bw[0]?('semaine du '+fmtJ(bw[0])):''],
  ].map(k=>`<div class="kpi"><div class="v">${k[1]}</div><div class="l">${k[0]}</div>`
    +(k[2]?`<div class="delta" style="color:var(--muted)">${k[2]}</div>`:'')+`</div>`).join('');
}

function renderEv(){
  const af=document.getElementById('evAct').value;
  const src=DATA.filter(r=>!af||r.activite===af);
  const m={};src.forEach(r=>{const k=periodKey(r,evGran);(m[k]=m[k]||[]).push(r);});
  const keys=Object.keys(m).sort();
  let vals, tip, ymax, line=false;
  if(evMetric==='taux'){
    vals=keys.map(k=>avg(m[k])); ymax=100; line=true;
    tip=c=>c.parsed.y.toFixed(1)+'%';
  }else if(evMetric==='seances'){
    vals=keys.map(k=>m[k].length);
    tip=c=>nf(c.parsed.y)+' séance(s)';
  }else{
    vals=keys.map(k=>m[k].reduce((s,r)=>s+r.presents,0));
    tip=c=>nf(c.parsed.y)+' visiteurs';
  }
  if(evChart)evChart.destroy();
  evChart=new Chart(cDay,{type:line?'line':'bar',data:{labels:keys.map(k=>labelPeriode(k,evGran)),
    datasets:[{data:vals,borderColor:'#d98b63',backgroundColor:line?'rgba(217,139,99,.15)':'#d98b63',
      fill:line,tension:.3,pointRadius:evGran==='jour'?1:3}]},
    options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:tip}}},
      scales:{y:{beginAtZero:true,max:ymax,ticks:{callback:v=>evMetric==='taux'?v+'%':nf(v)}}}}});
}

const sumP=arr=>arr.reduce((s,r)=>s+r.presents,0);
const moyP=arr=>arr.length?sumP(arr)/arr.length:0;

function renderAct(){
  const by=group(r=>r.activite); const labels=Object.keys(by).sort((a,b)=>sumP(by[b])-sumP(by[a]));
  if(actChart)actChart.destroy();
  actChart=new Chart(cAct,{type:'bar',data:{labels,
    datasets:[{data:labels.map(a=>sumP(by[a])),backgroundColor:'#d98b63'}]},
    options:{indexAxis:'y',plugins:{legend:{display:false},
      tooltip:{callbacks:{label:c=>nf(c.parsed.x)+' visiteurs'}}},
      scales:{x:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}

function renderHour(){
  const by=group(r=>r.heure); const hours=Object.keys(by).sort();
  if(hourChart)hourChart.destroy();
  hourChart=new Chart(cHour,{type:'bar',data:{labels:hours,
    datasets:[{data:hours.map(h=>sumP(by[h])),backgroundColor:'#d98b63'}]},
    options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>nf(c.parsed.y)+' visiteurs'}}},
      scales:{y:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}

function renderCoach(){
  const by=group(r=>r.coach||'(sans coach)');
  const list=Object.keys(by).filter(c=>by[c].length>=3).sort((a,b)=>moyP(by[b])-moyP(by[a]));
  if(coachChart)coachChart.destroy();
  coachChart=new Chart(cCoach,{type:'bar',data:{labels:list.map(c=>`${c} (${by[c].length} séances)`),
    datasets:[{data:list.map(c=>moyP(by[c])),backgroundColor:'#d98b63'}]},
    options:{indexAxis:'y',plugins:{legend:{display:false},
      tooltip:{callbacks:{label:c=>c.parsed.x.toFixed(1)+' visiteurs / séance'}}},
      scales:{x:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}

function renderCA(){
  const prix=parseFloat(document.getElementById('prix').value)||0;
  try{localStorage.setItem('reset_prix',prix);}catch(e){}
  const sumBy=g=>{const m={};DATA.forEach(r=>{const k=periodKey(r,g);m[k]=(m[k]||0)+r.presents;});return m;};
  const nbMois=new Set(DATA.map(r=>r.date.slice(0,7))).size||1;
  const nbSem=new Set(DATA.map(r=>lundi(r.date))).size||1;
  const nbJours=new Set(DATA.map(r=>r.date)).size||1;
  const totalCA=DATA.reduce((s,r)=>s+r.presents,0)*prix;
  document.getElementById('caKpis').innerHTML=[
    ['CA total estimé',eur(totalCA)],
    ['CA / mois (moy.)',eur(totalCA/nbMois)],
    ['CA / semaine (moy.)',eur(totalCA/nbSem)],
    ['CA / jour (moy.)',eur(totalCA/nbJours)],
  ].map(k=>`<div class="kpi"><div class="v">${k[1]}</div><div class="l">${k[0]}</div></div>`).join('');
  const m=sumBy(caGran), keys=Object.keys(m).sort();
  if(caChart)caChart.destroy();
  caChart=new Chart(cCA,{type:'bar',data:{labels:keys.map(k=>labelPeriode(k,caGran)),
    datasets:[{data:keys.map(k=>m[k]*prix),backgroundColor:'#d98b63'}]},
    options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>eur(c.parsed.y)}}},
      scales:{y:{ticks:{callback:v=>v.toLocaleString('fr-FR')+' €'}}}}});
}

const selAct=document.getElementById('fAct'), selDay=document.getElementById('fDay'), selEvAct=document.getElementById('evAct');
function populateFilters(){
  const fa=selAct.value, fd=selDay.value, ea=selEvAct.value;
  const acts=[...new Set(DATA.map(r=>r.activite))].sort();
  const days=JOURS.filter(j=>DATA.some(r=>r.jour===j));
  const optAll='<option value="">Toutes activités</option>'+acts.map(a=>`<option>${a}</option>`).join('');
  selAct.innerHTML=optAll;
  selEvAct.innerHTML=optAll;
  selDay.innerHTML='<option value="">Tous les jours</option>'+days.map(d=>`<option>${d}</option>`).join('');
  selAct.value=fa; selDay.value=fd; selEvAct.value=ea;
}

const cols=[['date','Date'],['jour','Jour'],['heure','Heure'],['activite','Activité'],
  ['coach','Coach'],['remplissage','Visiteurs / places'],['complet','Complet']];
document.querySelector('#tbl thead').innerHTML='<tr>'+cols.map((c,i)=>`<th data-i="${i}">${c[1]}</th>`).join('')+'</tr>';
function renderTable(){
  const q=document.getElementById('q').value.toLowerCase();
  const fa=selAct.value, fd=selDay.value;
  let rows=DATA.filter(r=>(!fa||r.activite===fa)&&(!fd||r.jour===fd)&&
    (!q||(r.activite+' '+r.coach+' '+r.date+' '+r.heure).toLowerCase().includes(q)));
  rows.sort((a,b)=>{let x=a[sortKey],y=b[sortKey];
    if(sortKey==='remplissage'){x=a.presents;y=b.presents;} return x>y?sortDir:x<y?-sortDir:0;});
  currentRows=rows;
  document.querySelector('#tbl tbody').innerHTML=rows.map(r=>{
    const w=r.capacite?Math.min(100*r.presents/r.capacite,100):0;
    return `<tr><td>${fmtJ(r.date)}</td><td>${r.jour}</td><td>${r.heure}</td><td>${r.activite}</td>
      <td>${r.coach}</td>
      <td><span class="bar"><span style="width:${w}%;background:var(--accent)"></span></span>${r.remplissage}</td>
      <td>${r.complet}</td></tr>`;}).join('');
}

function maxDate(){return DATA.reduce((m,r)=>r.date>m?r.date:m,'');}
function updateLabel(){
  const last=maxDate();
  document.getElementById('lastScrap').innerHTML = last
    ? `Données jusqu'au <b>${fmtJ(last)}</b> &middot; ${DATA.length} séances`
    : 'Aucune donnée';
}

function dayShift(iso,n){const d=new Date(iso+'T00:00:00');d.setDate(d.getDate()+n);return d.toISOString().slice(0,10);}
function sumRange(a,b){return DATA.reduce((s,r)=>(r.date>=a&&r.date<=b)?s+r.presents:s,0);}
function rng(a,b){return fmtJM(a)+' au '+fmtJM(b);}
function trendCard(title,a,b,pa,pb,start){
  const cur=sumRange(a,b);
  const hasPrev = pa>=start;            // periode precedente entierement couverte par les donnees
  const prev=hasPrev?sumRange(pa,pb):0;
  let delta=`<div class="delta" style="color:var(--muted)">période de comparaison hors historique</div>`;
  if(hasPrev && prev>0){const pct=(cur-prev)/prev*100, up=pct>=0;
    delta=`<div class="delta" style="color:${up?'var(--green)':'var(--red)'}">`
      +`${up?'▲':'▼'} ${Math.abs(pct).toFixed(0)} % vs ${rng(pa,pb)} (${nf(prev)} visiteurs)</div>`;}
  return `<div class="kpi"><div class="v">${nf(cur)} visiteurs</div>`
    +`<div class="l">${title} &middot; du ${rng(a,b)}</div>${delta}</div>`;
}
function renderTrend(){
  const end=maxDate(); const box=document.getElementById('trend');
  if(!end){box.innerHTML='';return;}
  const start=DATA.reduce((m,r)=>r.date<m?r.date:m,end);
  box.innerHTML=
    trendCard('7 derniers jours', dayShift(end,-6), end, dayShift(end,-13), dayShift(end,-7), start)+
    trendCard('30 derniers jours', dayShift(end,-29), end, dayShift(end,-59), dayShift(end,-30), start);
}
function rankList(id,entries){
  const mx=entries.length?entries[0][1]:0;
  document.getElementById(id).innerHTML=entries.map(([lbl,v])=>
    `<div class="rk"><span class="lbl">${lbl}</span>`
    +`<span class="track"><span style="width:${mx?Math.round(100*v/mx):0}%"></span></span>`
    +`<span class="val">${nf(v)}</span></div>`).join('');
}
function renderTop(){
  const byH={};DATA.forEach(r=>byH[r.heure]=(byH[r.heure]||0)+r.presents);
  rankList('topHour',Object.entries(byH).sort((a,b)=>b[1]-a[1]).slice(0,8));
  const byD={};DATA.forEach(r=>byD[r.jour]=(byD[r.jour]||0)+r.presents);
  rankList('topDay',JOURS.filter(j=>byD[j]).map(j=>[j,byD[j]]).sort((a,b)=>b[1]-a[1]));
}
function renderAll(){
  populateFilters();
  renderKpis();renderTrend();renderEv();renderTop();renderAct();renderHour();renderCoach();renderCA();
  renderTable();updateLabel();
}

// ---- interactions ----
document.querySelectorAll('#evSeg button').forEach(b=>b.onclick=()=>{
  evGran=b.dataset.g;document.querySelectorAll('#evSeg button').forEach(x=>x.classList.remove('on'));
  b.classList.add('on');renderEv();});
document.querySelectorAll('#metSeg button').forEach(b=>b.onclick=()=>{
  evMetric=b.dataset.m;document.querySelectorAll('#metSeg button').forEach(x=>x.classList.remove('on'));
  b.classList.add('on');renderEv();});
selEvAct.addEventListener('change',renderEv);
document.querySelectorAll('#caSeg button').forEach(b=>b.onclick=()=>{
  caGran=b.dataset.g;document.querySelectorAll('#caSeg button').forEach(x=>x.classList.remove('on'));
  b.classList.add('on');renderCA();});
document.querySelectorAll('#tbl thead th').forEach(th=>th.onclick=()=>{
  const k=cols[+th.dataset.i][0]; sortDir=(sortKey===k)?-sortDir:1; sortKey=k; renderTable();});
['q','fAct','fDay'].forEach(id=>document.getElementById(id).addEventListener('input',renderTable));
document.getElementById('prix').addEventListener('input',renderCA);

const sp=(()=>{try{return localStorage.getItem('reset_prix');}catch(e){return null;}})();
if(sp)document.getElementById('prix').value=sp;

// ---- Export Excel (CSV ; + BOM) ----
document.getElementById('btnExport').addEventListener('click',()=>{
  const c2=[['date','Date'],['jour','Jour'],['heure','Heure'],['activite','Activité'],['coach','Coach'],
    ['presents','Visiteurs'],['capacite','Places'],['complet','Complet']];
  const esc=v=>{v=(''+v).replace(/"/g,'""');return /[";\n]/.test(v)?`"${v}"`:v;};
  const lines=[c2.map(c=>c[1]).join(';')];
  currentRows.forEach(r=>lines.push(c2.map(c=>esc(c[0]==='date'?fmtJ(r.date):r[c[0]])).join(';')));
  const blob=new Blob(['﻿'+lines.join('\r\n')],{type:'text/csv;charset=utf-8;'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);
  a.download='reset_seances.csv';document.body.appendChild(a);a.click();a.remove();
});

// ---- Mise a jour incrementale (depuis le navigateur) ----
function buildRow(o){
  const ds=(o.date_start||'').slice(0,16).replace('T',' ');
  const date=ds.slice(0,10), heure=ds.slice(11,16);
  let jour=''; const d=new Date(date+'T00:00:00'); if(!isNaN(d.getTime())) jour=JOURS[(d.getDay()+6)%7];
  const p=o.validated_booking_count||0, c=o.effectif||0;
  return {date,jour,heure,activite:o.activity_name||'',coach:COACHES[o.coach]||(o.coach||''),
    presents:p,capacite:c,remplissage:p+'/'+c,'taux_%':c?Math.round(1000*p/c)/10:'',
    complet:o.full?'oui':'non',session_id:o.id};
}
async function fetchOffers(min,max){
  let page=1, all=[];
  while(true){
    const u=`${API}/book/v1/offer/?company=${COMPANY}&only_future_strict=false`
      +`&min_date=${min}&max_date=${max}&page_size=300&page=${page}`;
    const res=await fetch(u);
    if(!res.ok) throw new Error('HTTP '+res.status);
    const j=await res.json();
    all=all.concat(j.results||[]);
    if(!j.next_page || !(j.results||[]).length) break;
    page++;
  }
  return all;
}
const msg=document.getElementById('updMsg');
document.getElementById('btnUpdate').addEventListener('click',async()=>{
  const btn=document.getElementById('btnUpdate');
  const y=new Date(); y.setDate(y.getDate()-1);
  const yIso=`${y.getFullYear()}-${String(y.getMonth()+1).padStart(2,'0')}-${String(y.getDate()).padStart(2,'0')}`;
  const from=maxDate()||yIso;
  if(from>yIso){ msg.textContent='Déjà à jour ✓'; return; }
  btn.disabled=true; btn.textContent='Mise à jour…'; msg.textContent='';
  try{
    const offers=await fetchOffers(from,yIso);
    const map=new Map(DATA.map(r=>[r.session_id,r]));
    let added=0;
    offers.forEach(o=>{const r=buildRow(o); if(!map.has(r.session_id))added++; map.set(r.session_id,r);});
    DATA=[...map.values()].sort((a,b)=>(a.date+a.heure)>(b.date+b.heure)?1:-1);
    persist(); renderAll();
    msg.textContent = added? `${added} nouvelle(s) séance(s) jusqu'au ${fmtJ(yIso)} ✓` : 'À jour, rien de nouveau ✓';
  }catch(e){
    msg.textContent='Erreur : '+e.message+' (réessaie plus tard)';
  }finally{
    btn.disabled=false; btn.textContent='Mettre à jour';
  }
});

renderAll();
</script>
</body>
</html>"""


def main():
    ap = argparse.ArgumentParser(description="Scrape taux de remplissage des seances Re-SET")
    ap.add_argument("--start", default="2026-03-22", help="date de debut AAAA-MM-JJ (defaut: ouverture 22/03/2026)")
    ap.add_argument("--end", default=None, help="date de fin AAAA-MM-JJ (defaut: hier)")
    ap.add_argument("--csv", default="reset_seances.csv", help="fichier CSV de sortie")
    ap.add_argument("--xlsx", default="reset_seances.xlsx", help="fichier Excel de sortie")
    ap.add_argument("--html", default="index.html", help="dashboard HTML de sortie")
    args = ap.parse_args()

    start = dt.date.fromisoformat(args.start)
    end = dt.date.fromisoformat(args.end) if args.end else dt.date.today() - dt.timedelta(days=1)

    print(f"Recuperation des seances du {start} au {end} ...")
    coaches = fetch_coaches()
    offers = fetch_offers(start, end)
    print(f"{len(offers)} offres recuperees.")
    rows = build_rows(offers, coaches)

    write_csv(rows, args.csv)
    write_xlsx(rows, args.xlsx)
    write_html(rows, coaches, args.html, start, end)

    if rows:
        jours = sorted({r["date"] for r in rows})
        tot_p = sum(r["presents"] for r in rows)
        tot_c = sum(r["capacite"] for r in rows)
        moy = round(100 * tot_p / tot_c, 1) if tot_c else 0
        print(f"OK: {len(rows)} seances sur {len(jours)} jours "
              f"({jours[0]} -> {jours[-1]}), {tot_p} presents / {tot_c} places, "
              f"taux pondere {moy}%.")
    else:
        print("Aucune seance trouvee sur cette periode.")


if __name__ == "__main__":
    main()
