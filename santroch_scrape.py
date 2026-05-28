#!/usr/bin/env python3
"""Scrape le nombre d'inscrits par seance pour Sant-Roch (plateforme Mariana Tek).

Particularite : Mariana Tek ne garde PAS l'historique (les seances passees
renvoient capacite=0). On scrape donc la fenetre a venir et on ACCUMULE les
releves dans santroch_data.json (versionne) pour batir l'historique jour
apres jour. Inscrits = capacite - places disponibles.

Genere : santroch_seances.csv / .xlsx / santroch.html

Usage:
    python3 santroch_scrape.py            # scrape aujourd'hui -> +28 jours, merge, regenere
    python3 santroch_scrape.py --ahead 40
"""
import argparse
import csv
import datetime as dt
import json
import os
import sys
import time
import urllib.parse
import urllib.request

BASE = "https://santroch.marianatek.com/api/customer/v1"
REGION = "48541"
LOCATION = "48717"
STORE = "santroch_data.json"
JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def _get(url, retries=4):
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(2 ** attempt)
    raise RuntimeError(f"echec requete {url}: {last}")


def fetch_window(start, end):
    """Toutes les classes entre start et end (dates incluses)."""
    params = urllib.parse.urlencode({
        "min_start_date": start.isoformat(),
        "max_start_date": end.isoformat(),
        "page_size": 500,
        "location": LOCATION,
        "region": REGION,
    })
    url = f"{BASE}/classes?{params}"
    out = []
    while url:
        data = _get(url)
        out.extend(data.get("results", []))
        url = data.get("next")
    return out


def snapshot_row(c):
    """Releve d'une classe -> dict, ou None si inexploitable."""
    if c.get("is_cancelled"):
        return None
    cap = c.get("capacity") or 0
    av = c.get("available_spot_count")
    if not cap or av is None:
        return None
    start = c.get("start_datetime", "") or ""
    date_part, heure = start[:10], start[11:16]
    try:
        jour = JOURS_FR[dt.date.fromisoformat(date_part).weekday()]
    except ValueError:
        return None
    inscrits = max(cap - av, 0)
    ct = c.get("class_type") or {}
    instr = ", ".join(i.get("name", "") for i in (c.get("instructors") or [])) or ""
    return {
        "date": date_part,
        "jour": jour,
        "heure": heure,
        "activite": ct.get("name", "") or "",
        "coach": instr,
        "inscrits": inscrits,
        "capacite": cap,
        "remplissage": f"{inscrits}/{cap}",
        "complet": "oui" if av == 0 else "non",
        "session_id": str(c.get("id")),
        "releve": dt.date.today().isoformat(),
    }


def load_store():
    if os.path.exists(STORE):
        try:
            with open(STORE, encoding="utf-8") as f:
                return json.load(f)
        except (ValueError, OSError):
            pass
    return {}


def save_store(store):
    with open(STORE, "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=0, sort_keys=True)


FIELDS = ["date", "jour", "heure", "activite", "coach", "inscrits",
          "capacite", "remplissage", "complet", "session_id", "releve"]


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    print(f"-> {path}")


def write_xlsx(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (Excel ignore: pip install openpyxl)", file=sys.stderr)
        return
    wb = Workbook(); ws = wb.active; ws.title = "Seances"
    ws.append([h.capitalize() for h in FIELDS])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1f2d3a"); c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    dc = FIELDS.index("date") + 1
    for r in rows:
        ws.append([r[h] for h in FIELDS])
    for row in ws.iter_rows(min_row=2, min_col=dc, max_col=dc):
        try:
            row[0].value = dt.date.fromisoformat(row[0].value); row[0].number_format = "DD/MM/YYYY"
        except (ValueError, TypeError):
            pass
    for i, wdt in enumerate([12, 11, 7, 26, 22, 9, 9, 12, 9, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    wb.save(path); print(f"-> {path}")


def write_html(rows, path):
    chartjs = ""
    vendor = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor_chartjs.min.js")
    if os.path.exists(vendor):
        with open(vendor, encoding="utf-8") as f:
            chartjs = f.read()
    dates = sorted({r["date"] for r in rows})
    periode = (f"{dt.date.fromisoformat(dates[0]).strftime('%d/%m/%Y')} au "
               f"{dt.date.fromisoformat(dates[-1]).strftime('%d/%m/%Y')}") if dates else "—"
    html = (HTML_TEMPLATE
            .replace("__CHARTJS__", chartjs)
            .replace("__DATA__", json.dumps(rows, ensure_ascii=False))
            .replace("__GENERATED__", dt.datetime.now().strftime("%d/%m/%Y %H:%M"))
            .replace("__PERIODE__", periode))
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"-> {path}")


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sant-Roch - Fréquentation des séances</title>
<script>__CHARTJS__</script>
<style>
  :root{--bg:#0f1720;--card:#16202b;--card2:#1d2a38;--line:#2a3a4a;
        --text:#e6eef5;--muted:#90a4b6;--accent:#4db6c4;--accent2:#7fd3df;
        --green:#6fcf97;--yellow:#e6c14d;--red:#e07a6f;}
  *{box-sizing:border-box;}
  body{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:var(--bg);color:var(--text);}
  header{padding:28px 32px 12px;}
  h1{margin:0;font-size:24px;font-weight:700;letter-spacing:.3px;}
  .sub{color:var(--muted);font-size:13px;margin-top:6px;}
  .wrap{padding:0 32px 48px;max-width:1280px;margin:0 auto;}
  .note{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:13px 18px;margin:18px 0 4px;color:var(--muted);font-size:13px;}
  .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin:22px 0;}
  .kpi{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:18px 20px;}
  .kpi .v{font-size:28px;font-weight:700;color:var(--accent2);}
  .kpi .l{color:var(--muted);font-size:12px;margin-top:4px;text-transform:uppercase;letter-spacing:.6px;}
  .kpi .delta{margin-top:6px;font-size:13px;font-weight:700;}
  .grid{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px;}
  @media(max-width:880px){.grid{grid-template-columns:1fr;}}
  .panel{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:18px 20px;}
  .panel h2{margin:0 0 14px;font-size:14px;color:var(--accent2);font-weight:600;}
  canvas{max-height:260px;}
  .filters{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin:8px 0 16px;}
  .filters input,.filters select{background:var(--card2);color:var(--text);border:1px solid var(--line);border-radius:9px;padding:9px 12px;font-size:13px;}
  .filters input{flex:1;min-width:180px;}
  table{width:100%;border-collapse:collapse;font-size:13px;}
  th,td{padding:9px 10px;text-align:left;border-bottom:1px solid var(--line);white-space:nowrap;}
  th{color:var(--muted);font-weight:600;cursor:pointer;position:sticky;top:0;background:var(--card);}
  tbody tr:hover{background:var(--card2);}
  .bar{display:inline-block;height:7px;border-radius:4px;vertical-align:middle;margin-right:8px;width:60px;background:var(--line);position:relative;overflow:hidden;}
  .bar>span{position:absolute;left:0;top:0;bottom:0;border-radius:4px;}
  .ranklist{display:flex;flex-direction:column;gap:9px;}
  .rk{display:grid;grid-template-columns:62px 1fr auto;align-items:center;gap:10px;font-size:13px;}
  .rk .lbl{font-weight:600;}
  .rk .track{height:9px;background:var(--line);border-radius:5px;overflow:hidden;}
  .rk .track>span{display:block;height:100%;background:var(--accent);border-radius:5px;}
  .rk .val{color:var(--muted);}
  .seg button{background:var(--card2);color:var(--muted);border:1px solid var(--line);padding:8px 16px;font-size:13px;cursor:pointer;}
  .seg button:first-child{border-radius:9px 0 0 9px;} .seg button:last-child{border-radius:0 9px 9px 0;}
  .seg button:not(:last-child){border-right:none;} .seg button.on{background:var(--accent);color:#0f1720;font-weight:700;}
  .btn{background:var(--accent);color:#0f1720;border:none;border-radius:9px;padding:9px 16px;font-size:13px;font-weight:700;cursor:pointer;}
  .tablewrap{max-height:560px;overflow:auto;border:1px solid var(--line);border-radius:14px;}
  .foot{color:var(--muted);font-size:12px;margin-top:18px;}
  @media(max-width:600px){
    header{padding:18px 14px 6px;} h1{font-size:18px;} .wrap{padding:0 12px 32px;}
    .kpis{grid-template-columns:1fr 1fr;gap:10px;} .kpi .v{font-size:21px;}
    .seg{display:flex;width:100%;} .seg button{flex:1;padding:10px 4px;}
    .filters input,.filters select,.btn{font-size:15px;width:100%;}
  }
</style>
</head>
<body>
<header>
  <h1>Sant-Roch &middot; Fréquentation des séances</h1>
  <div class="sub">Période __PERIODE__ &middot; généré le __GENERATED__</div>
</header>
<div class="wrap">
  <div class="note" id="note"></div>
  <div class="kpis" id="kpis"></div>
  <div class="panel">
    <h2>Tendance de la fréquentation</h2>
    <div class="kpis" id="trend"></div>
  </div>
  <div class="panel">
    <h2>Fréquentation (inscrits)</h2>
    <div class="filters">
      <span class="seg" id="metSeg">
        <button data-m="inscrits" class="on">Inscrits</button>
        <button data-m="seances">Nb de séances</button>
      </span>
      <span class="seg" id="evSeg">
        <button data-g="jour" class="on">Jour</button>
        <button data-g="semaine">Semaine</button>
        <button data-g="mois">Mois</button>
      </span>
      <select id="evAct"></select>
    </div>
    <canvas id="cEv" style="margin-top:6px"></canvas>
  </div>
  <div class="grid">
    <div class="panel"><h2>Top créneaux horaires (inscrits)</h2><div id="topHour" class="ranklist"></div></div>
    <div class="panel"><h2>Inscrits par jour de la semaine</h2><div id="topDay" class="ranklist"></div></div>
  </div>
  <div class="grid">
    <div class="panel"><h2>Inscrits par type de rituel</h2><canvas id="cAct"></canvas></div>
    <div class="panel"><h2>Inscrits moyens par séance &amp; par guide</h2><canvas id="cCoach" style="max-height:none"></canvas></div>
  </div>
  <div class="panel">
    <h2>Détail des séances</h2>
    <div class="filters">
      <input id="q" placeholder="Rechercher (rituel, guide, date...)">
      <select id="fAct"></select>
      <select id="fDay"></select>
      <button id="btnExport" class="btn">Exporter en Excel</button>
    </div>
    <div class="tablewrap"><table id="tbl"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="foot">Inscrits = réservations (capacité − places disponibles). Source : sant-roch.com (Mariana Tek). Historique accumulé jour après jour.</div>
</div>
<script>
const DATA=__DATA__;
const JOURS=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"];
const MOIS=['janv.','févr.','mars','avr.','mai','juin','juil.','août','sept.','oct.','nov.','déc.'];
const col=t=>t>=85?'#6fcf97':t>=50?'#e6c14d':'#e07a6f';
const nf=v=>Math.round(v).toLocaleString('fr-FR');
Chart.defaults.color='#90a4b6';Chart.defaults.borderColor='#2a3a4a';
Chart.defaults.font.family=getComputedStyle(document.body).fontFamily;
function group(key){const m={};DATA.forEach(r=>{(m[key(r)]=m[key(r)]||[]).push(r);});return m;}
function lundi(s){const d=new Date(s+'T00:00:00');const j=(d.getDay()+6)%7;d.setDate(d.getDate()-j);return d.toISOString().slice(0,10);}
function periodKey(r,g){return g==='jour'?r.date:g==='mois'?r.date.slice(0,7):lundi(r.date);}
function fmtJ(iso){const p=iso.split('-');return `${p[2]}/${p[1]}/${p[0]}`;}
function fmtJM(iso){const p=iso.split('-');return `${p[2]}/${p[1]}`;}
function fmtMois(ym){const p=ym.split('-');return `${MOIS[+p[1]-1]} ${p[0]}`;}
function labelPeriode(k,g){return g==='mois'?fmtMois(k):g==='semaine'?'sem. '+fmtJM(k):fmtJM(k);}
const sumI=a=>a.reduce((s,r)=>s+r.inscrits,0), moyI=a=>a.length?sumI(a)/a.length:0;
const maxDate=()=>DATA.reduce((m,r)=>r.date>m?r.date:m,'');
const minDate=()=>DATA.reduce((m,r)=>(m===''||r.date<m)?r.date:m,'');

let evChart=null,actChart=null,coachChart=null,evGran='jour',evMetric='inscrits',sortKey='date',sortDir=1,currentRows=[];

function renderNote(){
  const future=DATA.some(r=>r.date>=new Date().toISOString().slice(0,10));
  document.getElementById('note').innerHTML=
    "ℹ️ Mariana Tek ne fournit pas l'historique passé : les chiffres sont relevés à l'avance et l'historique se construit jour après jour. "
    +(DATA.length?`Couverture actuelle : ${new Set(DATA.map(r=>r.date)).size} jours.`:'Pas encore de données.');
}
function renderKpis(){
  const tot=sumI(DATA);
  const byD={};DATA.forEach(r=>byD[r.date]=(byD[r.date]||0)+r.inscrits);
  const bd=Object.entries(byD).sort((a,b)=>b[1]-a[1])[0]||['',0];
  const byW={};DATA.forEach(r=>{const k=lundi(r.date);byW[k]=(byW[k]||0)+r.inscrits;});
  const bw=Object.entries(byW).sort((a,b)=>b[1]-a[1])[0]||['',0];
  document.getElementById('kpis').innerHTML=[
    ['Inscrits (total)',nf(tot),''],
    ['Séances',nf(DATA.length),''],
    ['Meilleure journée',nf(bd[1]),bd[0]?fmtJ(bd[0]):''],
    ['Meilleure semaine',nf(bw[1]),bw[0]?('semaine du '+fmtJ(bw[0])):''],
  ].map(k=>`<div class="kpi"><div class="v">${k[1]}</div><div class="l">${k[0]}</div>`
    +(k[2]?`<div class="delta" style="color:var(--muted)">${k[2]}</div>`:'')+`</div>`).join('');
}
function dayShift(iso,n){const d=new Date(iso+'T00:00:00');d.setDate(d.getDate()+n);return d.toISOString().slice(0,10);}
function sumRange(a,b){return DATA.reduce((s,r)=>(r.date>=a&&r.date<=b)?s+r.inscrits:s,0);}
function trendCard(title,a,b,pa,pb,start){
  const cur=sumRange(a,b), hasPrev=pa>=start, prev=hasPrev?sumRange(pa,pb):0;
  let delta='<div class="delta" style="color:var(--muted)">période de comparaison hors couverture</div>';
  if(hasPrev&&prev>0){const pct=(cur-prev)/prev*100,up=pct>=0;
    delta=`<div class="delta" style="color:${up?'var(--green)':'var(--red)'}">${up?'▲':'▼'} ${Math.abs(pct).toFixed(0)} % vs ${fmtJM(pa)} au ${fmtJM(pb)} (${nf(prev)})</div>`;}
  return `<div class="kpi"><div class="v">${nf(cur)} inscrits</div><div class="l">${title} &middot; du ${fmtJM(a)} au ${fmtJM(b)}</div>${delta}</div>`;
}
function renderTrend(){
  const end=maxDate(),box=document.getElementById('trend');
  if(!end){box.innerHTML='';return;}
  const start=minDate();
  box.innerHTML=trendCard('7 derniers jours',dayShift(end,-6),end,dayShift(end,-13),dayShift(end,-7),start)
    +trendCard('30 derniers jours',dayShift(end,-29),end,dayShift(end,-59),dayShift(end,-30),start);
}
function renderEv(){
  const af=document.getElementById('evAct').value;
  const src=DATA.filter(r=>!af||r.activite===af);
  const m={};src.forEach(r=>{const k=periodKey(r,evGran);(m[k]=m[k]||[]).push(r);});
  const keys=Object.keys(m).sort();
  const vals=keys.map(k=>evMetric==='seances'?m[k].length:sumI(m[k]));
  const tip=evMetric==='seances'?(c=>nf(c.parsed.y)+' séance(s)'):(c=>nf(c.parsed.y)+' inscrits');
  if(evChart)evChart.destroy();
  evChart=new Chart(cEv,{type:'bar',data:{labels:keys.map(k=>labelPeriode(k,evGran)),
    datasets:[{data:vals,backgroundColor:'#4db6c4'}]},
    options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:tip}}},
      scales:{y:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}
function renderAct(){
  const by=group(r=>r.activite); const labels=Object.keys(by).sort((a,b)=>sumI(by[b])-sumI(by[a]));
  if(actChart)actChart.destroy();
  actChart=new Chart(cAct,{type:'bar',data:{labels,datasets:[{data:labels.map(a=>sumI(by[a])),backgroundColor:'#4db6c4'}]},
    options:{indexAxis:'y',plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>nf(c.parsed.x)+' inscrits'}}},
      scales:{x:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}
function renderCoach(){
  const by=group(r=>r.coach||'(sans guide)');
  const list=Object.keys(by).filter(c=>by[c].length>=3).sort((a,b)=>moyI(by[b])-moyI(by[a]));
  if(coachChart)coachChart.destroy();
  coachChart=new Chart(cCoach,{type:'bar',data:{labels:list.map(c=>`${c} (${by[c].length})`),
    datasets:[{data:list.map(c=>moyI(by[c])),backgroundColor:'#4db6c4'}]},
    options:{indexAxis:'y',plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.parsed.x.toFixed(1)+' inscrits / séance'}}},
      scales:{x:{beginAtZero:true,ticks:{callback:v=>nf(v)}}}}});
}
function rankList(id,entries){
  const mx=entries.length?entries[0][1]:0;
  document.getElementById(id).innerHTML=entries.map(([lbl,v])=>
    `<div class="rk"><span class="lbl">${lbl}</span><span class="track"><span style="width:${mx?Math.round(100*v/mx):0}%"></span></span><span class="val">${nf(v)}</span></div>`).join('');
}
function renderTop(){
  const byH={};DATA.forEach(r=>byH[r.heure]=(byH[r.heure]||0)+r.inscrits);
  rankList('topHour',Object.entries(byH).sort((a,b)=>b[1]-a[1]).slice(0,8));
  const byD={};DATA.forEach(r=>byD[r.jour]=(byD[r.jour]||0)+r.inscrits);
  rankList('topDay',JOURS.filter(j=>byD[j]).map(j=>[j,byD[j]]).sort((a,b)=>b[1]-a[1]));
}
const selAct=document.getElementById('fAct'),selDay=document.getElementById('fDay'),selEvAct=document.getElementById('evAct');
function populateFilters(){
  const fa=selAct.value,fd=selDay.value,ea=selEvAct.value;
  const acts=[...new Set(DATA.map(r=>r.activite))].sort();
  const days=JOURS.filter(j=>DATA.some(r=>r.jour===j));
  const all='<option value="">Tous les rituels</option>'+acts.map(a=>`<option>${a}</option>`).join('');
  selAct.innerHTML=all;selEvAct.innerHTML=all;
  selDay.innerHTML='<option value="">Tous les jours</option>'+days.map(d=>`<option>${d}</option>`).join('');
  selAct.value=fa;selDay.value=fd;selEvAct.value=ea;
}
const cols=[['date','Date'],['jour','Jour'],['heure','Heure'],['activite','Rituel'],['coach','Guide'],['remplissage','Inscrits / cap.'],['complet','Complet']];
document.querySelector('#tbl thead').innerHTML='<tr>'+cols.map((c,i)=>`<th data-i="${i}">${c[1]}</th>`).join('')+'</tr>';
function renderTable(){
  const q=document.getElementById('q').value.toLowerCase(),fa=selAct.value,fd=selDay.value;
  let rows=DATA.filter(r=>(!fa||r.activite===fa)&&(!fd||r.jour===fd)&&(!q||(r.activite+' '+r.coach+' '+r.date+' '+r.heure).toLowerCase().includes(q)));
  rows.sort((a,b)=>{let x=a[sortKey],y=b[sortKey];if(sortKey==='remplissage'){x=a.inscrits;y=b.inscrits;}return x>y?sortDir:x<y?-sortDir:0;});
  currentRows=rows;
  document.querySelector('#tbl tbody').innerHTML=rows.map(r=>{
    const ratio=r.capacite?100*r.inscrits/r.capacite:0;
    return `<tr><td>${fmtJ(r.date)}</td><td>${r.jour}</td><td>${r.heure}</td><td>${r.activite}</td><td>${r.coach}</td>`
      +`<td><span class="bar"><span style="width:${Math.min(ratio,100)}%;background:${col(ratio)}"></span></span>${r.remplissage}</td><td>${r.complet}</td></tr>`;}).join('');
}
function renderAll(){populateFilters();renderNote();renderKpis();renderTrend();renderEv();renderTop();renderAct();renderCoach();renderTable();}
document.querySelectorAll('#evSeg button').forEach(b=>b.onclick=()=>{evGran=b.dataset.g;document.querySelectorAll('#evSeg button').forEach(x=>x.classList.remove('on'));b.classList.add('on');renderEv();});
document.querySelectorAll('#metSeg button').forEach(b=>b.onclick=()=>{evMetric=b.dataset.m;document.querySelectorAll('#metSeg button').forEach(x=>x.classList.remove('on'));b.classList.add('on');renderEv();});
selEvAct.addEventListener('change',renderEv);
document.querySelectorAll('#tbl thead th').forEach(th=>th.onclick=()=>{const k=cols[+th.dataset.i][0];sortDir=(sortKey===k)?-sortDir:1;sortKey=k;renderTable();});
['q','fAct','fDay'].forEach(id=>document.getElementById(id).addEventListener('input',renderTable));
document.getElementById('btnExport').addEventListener('click',()=>{
  const c2=[['date','Date'],['jour','Jour'],['heure','Heure'],['activite','Rituel'],['coach','Guide'],['inscrits','Inscrits'],['capacite','Capacité'],['complet','Complet']];
  const esc=v=>{v=(''+v).replace(/"/g,'""');return /[";\n]/.test(v)?`"${v}"`:v;};
  const lines=[c2.map(c=>c[1]).join(';')];
  currentRows.forEach(r=>lines.push(c2.map(c=>esc(c[0]==='date'?fmtJ(r.date):r[c[0]])).join(';')));
  const blob=new Blob(['﻿'+lines.join('\r\n')],{type:'text/csv;charset=utf-8;'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='santroch_seances.csv';document.body.appendChild(a);a.click();a.remove();
});
renderAll();
</script>
</body>
</html>"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ahead", type=int, default=28, help="nb de jours a venir a scraper")
    args = ap.parse_args()

    today = dt.date.today()
    end = today + dt.timedelta(days=args.ahead)
    print(f"Scrap Sant-Roch du {today} au {end} ...")
    classes = fetch_window(today, end)
    print(f"{len(classes)} classes recuperees.")

    store = load_store()
    added = 0
    for c in classes:
        row = snapshot_row(c)
        if row:
            if row["session_id"] not in store:
                added += 1
            store[row["session_id"]] = row
    save_store(store)
    print(f"{added} nouvelles seances, {len(store)} au total en historique.")

    rows = sorted(store.values(), key=lambda r: (r["date"], r["heure"]))
    write_csv(rows, "santroch_seances.csv")
    write_xlsx(rows, "santroch_seances.xlsx")
    write_html(rows, "santroch.html")

    if rows:
        tot = sum(r["inscrits"] for r in rows)
        jours = sorted({r["date"] for r in rows})
        print(f"OK: {len(rows)} seances sur {len(jours)} jours "
              f"({jours[0]} -> {jours[-1]}), {tot} inscrits cumules.")


if __name__ == "__main__":
    main()
